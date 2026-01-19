"""
Data utilities for AutoSARM.

This module provides data processing and conversion functions.
"""

from __future__ import annotations

import logging
import os
import re
from pathlib import Path
from typing import List, Optional, Union

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image
from rdkit.Chem import Draw

from autosarm.utils.mol_utils import safe_mol_from_smiles

logger = logging.getLogger(__name__)


def get_number_from_string(text: str) -> Optional[float]:
    """
    Extract number from a string.
    
    Args:
        text: Input string potentially containing a number
        
    Returns:
        Extracted float or None if no number found
    """
    text = str(text)
    pattern = r'[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?'
    matches = re.findall(pattern, text)
    
    if matches:
        return float(matches[0])
    return None


def float_row(
    df: pd.DataFrame, 
    cols: List[str], 
    dropna: bool = True
) -> pd.DataFrame:
    """
    Convert specified columns to float type.
    
    Args:
        df: Input DataFrame
        cols: Column names to convert
        dropna: Whether to drop rows with invalid values
        
    Returns:
        DataFrame with converted columns
    """
    df = df.copy()
    
    for col in cols:
        if col not in df.columns:
            logger.warning(f"Column '{col}' not found in DataFrame")
            continue
            
        error_indices = []
        
        for idx, row in df.iterrows():
            data_val = df.loc[idx, col]
            try:
                data_val = get_number_from_string(data_val)
                value = float(data_val)
                df.loc[idx, col] = value
            except (TypeError, ValueError):
                error_indices.append(idx)
                df.loc[idx, col] = np.nan
        
        if dropna and error_indices:
            df = df.drop(labels=error_indices)
        
        df[col] = df[col].astype(float)
    
    return df


def load_csv_or_df(csv_or_df: Union[str, pd.DataFrame]) -> pd.DataFrame:
    """
    Load data from CSV file path or return DataFrame copy.
    
    Args:
        csv_or_df: CSV file path or DataFrame
        
    Returns:
        DataFrame
        
    Raises:
        TypeError: If input is neither string nor DataFrame
    """
    if isinstance(csv_or_df, str):
        return pd.read_csv(csv_or_df)
    elif isinstance(csv_or_df, pd.DataFrame):
        return csv_or_df.copy()
    else:
        raise TypeError("Input should be a file path string or DataFrame")


def csv_to_excel(
    csv_path: str,
    img_cols: List[str] = None,
    save_file: str = '',
    max_imgs: int = 500,
    column_as_image: bool = False
) -> openpyxl.Workbook:
    """
    Convert CSV file to Excel with molecular images.
    
    Args:
        csv_path: Path to input CSV file
        img_cols: Columns containing SMILES to render as images
        save_file: Output Excel file path
        max_imgs: Maximum number of images to generate
        column_as_image: Whether to render column headers as images
        
    Returns:
        OpenPyXL Workbook object
    """
    img_cols = img_cols or ['SMILES', 'smi']
    
    df = pd.read_csv(csv_path)
    letters = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Sheet1')
    
    tmp_img_path = Path('./images_png')
    tmp_img_path.mkdir(exist_ok=True, parents=True)
    
    col_names = df.columns.tolist()
    
    # Write column headers
    for icol, val in enumerate(col_names):
        irow = 1
        
        if column_as_image:
            try:
                col_letter = letters[icol]
                ws.row_dimensions[irow].height = 90
                ws.column_dimensions[col_letter].width = 20
                
                mol = safe_mol_from_smiles(val)
                if mol is not None:
                    img = Draw.MolToImage(mol, size=[200, 200])
                    img_path = tmp_img_path / f'molecule{irow}{icol}.png'
                    img.save(img_path)
                    
                    excel_img = Image(str(img_path))
                    excel_img.width = 100
                    excel_img.height = 100
                    excel_img.anchor = ws.cell(row=irow, column=icol + 1).coordinate
                    ws.add_image(excel_img)
            except Exception as e:
                logger.debug(f"Error rendering column header: {e}")
        
        ws.cell(irow, icol + 1).value = val
    
    # Write data rows
    for idx, row in df.iterrows():
        irow = idx + 2  # Row index starts from 1, and first row is header
        
        for icol, vcol in enumerate(col_names):
            if vcol in img_cols and icol <= max_imgs:
                try:
                    col_letter = letters[icol]
                    ws.row_dimensions[irow].height = 90
                    ws.column_dimensions[col_letter].width = 25
                    
                    mol = safe_mol_from_smiles(row[vcol])
                    if mol is not None:
                        img = Draw.MolToImage(mol, size=[400, 200])
                        img_path = tmp_img_path / f'molecule{irow}{icol}.png'
                        img.save(img_path)
                        
                        excel_img = Image(str(img_path))
                        excel_img.width = 200
                        excel_img.height = 100
                        excel_img.anchor = ws.cell(row=irow, column=icol + 1).coordinate
                        ws.add_image(excel_img)
                except Exception as e:
                    logger.debug(f"Error rendering cell image: {e}")
            
            ws.cell(irow, icol + 1).value = row[vcol]
    
    wb.save(save_file)
    wb.close()
    
    # Cleanup temporary images
    import shutil
    shutil.rmtree(tmp_img_path, ignore_errors=True)
    
    return wb


def df_valid(
    df: pd.DataFrame, 
    row_smi: str = 'smiles'
) -> pd.DataFrame:
    """
    Filter DataFrame to only include rows with valid SMILES.
    
    Args:
        df: Input DataFrame
        row_smi: Column name containing SMILES
        
    Returns:
        Filtered DataFrame with valid SMILES only
    """
    from autosarm.utils.mol_utils import get_mol
    
    df = df.copy()
    valid_mask = df[row_smi].apply(lambda x: get_mol(x) is not None)
    return df[valid_mask].reset_index(drop=True)


def sort_mol_sim_df(
    df: pd.DataFrame, 
    col: str = 'OrgSmi'
) -> pd.DataFrame:
    """
    Sort DataFrame by molecular similarity.
    
    Arranges molecules so that similar molecules are adjacent.
    
    Args:
        df: Input DataFrame with SMILES column
        col: Column name containing SMILES
        
    Returns:
        Reordered DataFrame
    """
    from autosarm.utils.mol_utils import get_mol, compute_fingerprint
    from rdkit import DataStructs
    
    df_sorted = pd.DataFrame(columns=['Index', 'Smi', 'Mol', 'Fp'])
    
    for idx, row in df.iterrows():
        smi = row[col]
        mol = get_mol(smi)
        fp = compute_fingerprint(mol)
        
        if len(df_sorted) == 0:
            df_sorted.loc[0] = [idx, smi, mol, fp]
            continue
        
        # Find most similar existing molecule
        sim_array = np.array([
            DataStructs.TanimotoSimilarity(fp, existing_fp) 
            for existing_fp in df_sorted['Fp']
        ])
        max_idx = np.argmax(sim_array)
        
        # Insert after most similar
        df_sorted.loc[max_idx + 0.5] = [idx, smi, mol, fp]
        df_sorted = df_sorted.sort_index().reset_index(drop=True)
    
    return df.loc[df_sorted["Index"]]
