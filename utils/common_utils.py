from rdkit import Chem, DataStructs
from rdkit.Chem import AllChem,Draw,rdFMCS
from rdkit.Chem import rdFingerprintGenerator
from functools import partial
import numpy as np
import pandas as pd
from multiprocessing import Pool
import copy
import openpyxl
from openpyxl.drawing.image import Image 
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
from pathlib import Path
import os,sys
import re
import logging

def is_valid_smiles(smi):
    """验证SMILES字符串是否有效"""
    # 类型检查
    if not isinstance(smi, str):
        return False
    
    # 过滤纯数字和空字符串
    smi_stripped = smi.strip()
    if not smi_stripped or smi_stripped.isdigit():
        return False
    
    # 尝试解析SMILES
    try:
        mol = Chem.MolFromSmiles(smi_stripped)
        return mol is not None
    except:
        return False

def safe_mol_from_smiles(smi, log_errors=False):
    """安全地从SMILES创建分子对象，带类型检查"""
    # 类型检查 - 处理numpy类型和Python原生类型
    if isinstance(smi, (int, float, np.integer, np.floating)):
        if log_errors:
            logging.warning(f"Received numeric value {smi} instead of SMILES string")
        return None
    
    # 转换为字符串
    smi_str = str(smi) if smi is not None else ""
    
    # 验证SMILES
    if not is_valid_smiles(smi_str):
        if log_errors:
            logging.warning(f"Invalid SMILES: {smi_str}")
        return None
    
    try:
        return Chem.MolFromSmiles(smi_str)
    except Exception as e:
        if log_errors:
            logging.error(f"Error parsing SMILES '{smi_str}': {e}")
        return None

def get_number_from_string(text):
    ''' get number from string  '''
    text=str(text)
    # pattern = r'\d+\.\d+'
    pattern = r'[+-]?(?:\d+\.?\d*|\.\d+)(?:[eE][+-]?\d+)?'
    matches = re.findall(pattern, text)
    if len(matches)>0:
        return float(matches[0])
    else:
        return None


def float_row(df, cols=[],dropna=True):
    ''' transform the row format into float '''
    for icol in cols:
        data_error_index=[]
        for idx,row in df.iterrows():
            data_col=df.loc[idx,icol]
            try:
                data_col=get_number_from_string(data_col)
                value=float(data_col)
                df.loc[idx,icol]=value
                # value=value>20
            except:
                data_error_index.append(idx)
                df.loc[idx,icol]=np.nan
        if dropna:
            df=df.drop(labels=data_error_index)
        df[icol]=df[icol].astype(float)
    return df

def heavy_atm_smiles(smi):
    ''' count heavy atoms '''
    heavy_atoms=['C','c','N','n','O','o','S','s','F','Br','Cl','I','B','P']
    count=0
    for isymbl in smi:
        if isymbl in heavy_atoms:
            count+=1
    return count

def kekulize_smi(smi):
    '''  kekulize the SMILES  '''
    # print("smi:  ", smi)
    try:
        mol=get_mol(smi)
        Chem.Kekulize(mol, clearAromaticFlags=True)
        smi_kekule=Chem.MolToSmiles(mol, kekuleSmiles=True)
        # print("kekulize_smi:  ", smi_kekule)
        return smi_kekule
    except Exception as e:
        print(e)
        return smi

def csvToExcel(csv, imgCols=['SMILES','smi'],save_file='',max_imgs=500,column=False):
    '''  Transform Pandas dataframe into Excel  '''
    df=pd.read_csv(csv)
    LETTERS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'  ## the maximum columns cannot too long
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Sheet1')
    tmpImgPath=Path('./images_png')
    tmpImgPath.mkdir(exist_ok=True, parents=True)
    col_names=df.columns.to_list()
    ###  Write column head
    for icol,val in enumerate(col_names):
        irow=1
        if column:
            try:
            # if 1:  ## for dubugging
                col_letter=LETTERS[icol]
                ws.row_dimensions[irow].height = 90
                ws.column_dimensions[col_letter].width = 20
                mol = safe_mol_from_smiles(val)
                if mol is None:
                    continue
                img = Draw.MolToImage(mol, size=[200, 200])
                img.save(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                img=Image(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                img.width = 100
                img.height = 100
                # print(ws.cell(row=irow, column=icol).coordinate)
                img.anchor = ws.cell(row=irow, column=icol+1).coordinate # col and row are zero-based
                # ws.add_image(img, f"{col_letter}{irow}")
                ws.add_image(img)
            except Exception as e:
                print(e)
                continue
        ws.cell(irow,icol+1).value=val
    for idx,row in df.iterrows():
        irow=idx+2 ## the row id start from one and the first was occupied by head
        for icol,vcol in enumerate(col_names):
            if vcol in imgCols:
                if icol>max_imgs:
                    continue
                try:
                # if 1:  ## for dubugging
                    col_letter=LETTERS[icol]
                    ws.row_dimensions[irow].height = 90
                    ws.column_dimensions[col_letter].width = 25
                    mol = safe_mol_from_smiles(row[vcol])
                    if mol is None:
                        continue
                    img = Draw.MolToImage(mol, size=[400, 200])
                    img.save(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                    img=Image(tmpImgPath.joinpath(f'molecule{irow}{icol}.png'))
                    img.width = 200
                    img.height = 100
                    # print(ws.cell(row=irow, column=icol).coordinate)
                    img.anchor = ws.cell(row=irow, column=icol+1).coordinate # col and row are zero-based
                    # ws.add_image(img, f"{col_letter}{irow}")
                    ws.add_image(img)
                except Exception as e:
                    print(e)
                    continue
            ws.cell(irow,icol+1).value=row[vcol]  ## add value to the folders
    wb.save(save_file)
    wb.close()
    os.system(f"rm -rf {tmpImgPath}")
    return wb

def get_mol(smiles_or_mol):
    '''
    Loads SMILES/molecule into RDKit's object
    '''
    if isinstance(smiles_or_mol, str):
        if len(smiles_or_mol) == 0:
            return None
        mol = Chem.MolFromSmiles(smiles_or_mol)
        if mol is None:
            return None
        try:
            Chem.SanitizeMol(mol)
        except ValueError:
            return None
        return mol
    return smiles_or_mol

def canonic_smiles(smiles_or_mol):
    try:
        mol = get_mol(smiles_or_mol)    
        if mol is None:
            return None
        return Chem.MolToSmiles(mol,isomericSmiles=False)
    except Exception as e:
        print(e)
        return None
    
# 创建全局MorganGenerator实例（提高性能）
_morgan_generator = None

def get_morgan_generator(radius=2, fpSize=1024):
    """获取或创建MorganGenerator实例"""
    global _morgan_generator
    if _morgan_generator is None:
        _morgan_generator = rdFingerprintGenerator.GetMorganGenerator(radius=radius, fpSize=fpSize)
    return _morgan_generator

def compute_FP(mol, radius=2, nBits=1024):
    mol = get_mol(mol)
    if mol is None:
        return None
    try:
        # 使用新的MorganGenerator API
        generator = get_morgan_generator(radius=radius, fpSize=nBits)
        FP = generator.GetFingerprint(mol)
        return FP
    except Exception as e:
        logging.warning(f"Error computing fingerprint: {e}")
        return None

def compute_sim(smi, smi_list, mode='smi-smi'):
    generator = get_morgan_generator(radius=2, fpSize=1024)
    
    if mode=='smi-smis':
        mol1 = safe_mol_from_smiles(smi)
        if mol1 is None:
            return [0] * len(smi_list)
        FP1 = generator.GetFingerprint(mol1)
        mols = [safe_mol_from_smiles(ismi) for ismi in smi_list]
        FPs = [generator.GetFingerprint(imol) for imol in mols if imol is not None]
        molSims = [DataStructs.TanimotoSimilarity(FP, FP1) for FP in FPs]
        
    elif mode=='smi-smi':
        mol1 = safe_mol_from_smiles(smi)
        mol2 = safe_mol_from_smiles(smi_list)
        if mol1 is None or mol2 is None:
            return 0

        FP1 = generator.GetFingerprint(mol1)
        FP2 = generator.GetFingerprint(mol2)
        molSims = DataStructs.TanimotoSimilarity(FP1, FP2)
        
    elif mode=='smi-FPs':
        mol1 = safe_mol_from_smiles(smi)
        if mol1 is None:
            return [0] * len(smi_list)
        FP1 = generator.GetFingerprint(mol1)
        FPs = smi_list
        molSims = [DataStructs.TanimotoSimilarity(FP, FP1) for FP in FPs]
    else:
        return 0
        
    return molSims

def mapper(n_jobs):
    '''
    Returns function for map call.
    If n_jobs == 1, will use standard map
    If n_jobs > 1, will use multiprocessing pool
    If n_jobs is a pool object, will return its map function
    '''
    if n_jobs == 1:
        def _mapper(*args, **kwargs):
            return list(map(*args, **kwargs))

        return _mapper
    if isinstance(n_jobs, int):
        pool = Pool(n_jobs)

        def _mapper(*args, **kwargs):
            try:
                result = pool.map(*args, **kwargs)
            finally:
                pool.terminate()
            return result

        return _mapper
    return n_jobs.map

def show_mols(smiles_mols,subImgSize=(500,400),legends=[]):
    '''Display multiple mols with legends'''
    mols=[get_mol(ismiles_mol) for ismiles_mol in smiles_mols]
    mol_cls,legends_cls=[],[]
    for i in range(len(mols)):
        if mols[i]==None:
            continue
        mol_cls.append(mols[i])
        if len(legends)>i:
            legends_cls.append(legends[i])
        else:
            legends_cls.append('')
    svg=Draw.MolsToGridImage(mol_cls,subImgSize=subImgSize, molsPerRow=3,useSVG=True,legends=legends_cls)
    png=Draw.MolsToGridImage(mol_cls,subImgSize=subImgSize,useSVG=False,molsPerRow=3,legends=legends_cls)
    return png,svg

def mol_with_atom_index(mol, Idx=None, start=1):
    mol=get_mol(mol)
    mol = copy.deepcopy(mol)
    for atom in mol.GetAtoms():
        if Idx != None:
            # atom.SetAtomMapNum(Idx)
            atom.SetProp("atomNote", str(Idx))
        else:
            atom.SetProp("atomNote", str(atom.GetIdx()+start))
    return mol

def sort_mol_sim_df(df, col='OrgSmi'):
    ''' Sort the dataframe based on the similarity of CPDs!  '''
    # df['index']=df[col]
    # df=df.set_index('index')
    df_sorted=pd.DataFrame.from_dict({'Index':[],'Smi':[],'Mol':[],'Fp':[]})
    for idx,row in df.iterrows():
        #for idx,ismi in enumerate(smi_list):
        ismi=row[col]
        mol=get_mol(ismi)
        Fp=compute_FP(mol)
        if len(df_sorted)==0:
            df_sorted.loc[0]=[idx, ismi, mol, Fp]
            continue
        Sim_np=np.array([DataStructs.TanimotoSimilarity(
                Fp,iFp) for iFp in df_sorted['Fp']])
        SimArgmax=np.argmax(Sim_np)
        df_sorted.loc[SimArgmax+0.5]=[idx,ismi, mol, Fp]
        df_sorted=df_sorted.sort_index()
        df_sorted=df_sorted.reset_index(drop=True)
    df_sorted=df.loc[df_sorted["Index"]]
    return df_sorted

def remove_dummy(smi):
    FragMol=Chem.MolFromSmiles(smi)
    matched=FragMol.GetSubstructMatches(Chem.MolFromSmarts("[#7][#0]"))
    atoms=FragMol.GetAtoms()
    for imatch in matched:
        atoms[imatch[1]].SetAtomicNum(1)
    # imol.UpdatePropertyCache()
    Chem.SanitizeMol(FragMol)
    FragMol=Chem.RemoveHs(FragMol)
    smi=Chem.MolToSmiles(FragMol)
    re_p=re.compile(r'\(\*\)|\*|\[\*\]|\-')
    ifrag_nodummy = re.sub(re_p, '', smi)  ## remove dummy atoms
    return ifrag_nodummy

def compute_sim_hierarchy(smi1,smi2):
    ''' compute similarity based on the fingerprint of the smaller fragement
    same(FP_bits)/smaller(FP_bits)
    '''
    FP1=compute_FP(remove_dummy(smi1))
    FP2=compute_FP(remove_dummy(smi2))

    np_fp1 = np.zeros((1,))
    DataStructs.ConvertToNumpyArray(FP1, np_fp1)

    np_fp2 = np.zeros((1,))
    DataStructs.ConvertToNumpyArray(FP2, np_fp2)

    minLen=np.array([np_fp1.sum(),np_fp2.sum()]).min()
    sameBits=(np_fp1*np_fp2).sum()
    simHierarchy=round(sameBits/minLen,2)
    return simHierarchy

def sort_mol_hierarchy_df(df, col='OrgSmi'):
    ''' Sort the dataframe based on the similarity of CPDs!  '''
    # df['index']=df[col]
    # df=df.set_index('index')
    df_sorted=pd.DataFrame.from_dict({'Index':[],'Smi':[]})
    for idx,row in df.iterrows():
        #for idx,ismi in enumerate(smi_list):
        ismi=row[col]
        # mol=get_mol(ismi)
        # Fp=compute_FP(mol)
        if len(df_sorted)==0:
            df_sorted.loc[0]=[idx, ismi]
            continue
        Sim_np=np.array([compute_sim_hierarchy(
                ismi,jSmi) for jSmi in df_sorted['Smi']])
        SimArgmax=np.argmax(Sim_np)
        df_sorted.loc[SimArgmax+0.5]=[idx,ismi]
        df_sorted=df_sorted.sort_index()
        df_sorted=df_sorted.reset_index(drop=True)
    df_sorted=df.loc[df_sorted["Index"]]
    return df_sorted
