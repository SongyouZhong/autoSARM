'''  create sar table from core and activity database  '''
from rdkit import Chem, DataStructs
from rdkit.Chem import AllChem,Draw
from rdkit.Chem import MCS
import numpy as np
import pandas as pd
import copy,re
from pathlib import Path
from utils.common_utils import mapper,get_mol,compute_FP,mol_with_atom_index,get_number_from_string, compute_sim
from utils.sarm_utils import has_match
from functools import partial
from functools import reduce
from pandarallel import pandarallel
from my_toolset.drawing_utils import show_mols
from my_toolset.my_utils import canonic_smiles
import re

from collections import Counter
import openpyxl
from openpyxl.drawing.image import Image 
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker
import re,copy
import shutil
import os,sys


import warnings
from rdkit import RDLogger
logger = RDLogger.logger()
logger.setLevel(RDLogger.ERROR)


def replace_pattern(text):
    """
    Replace all occurrences of "[1*]" with "[*:1]" in the input text.
    
    Parameters:
    text (str): Input string containing the pattern to replace
    
    Returns:
    str: Modified string with replacements
    """
    # Use regular expression to find and replace the pattern
    # The \[ and \] escape the square brackets since they have special meaning in regex
    modified_text = re.sub(r"\[1\*\]", r"[*:1]", text)
    modified_text = re.sub(r"\[2\*\]", r"[*:2]", modified_text)
    modified_text = re.sub(r"\[3\*\]", r"[*:3]", modified_text)

    # modified_text = re.sub(rf"\[1\*\]", rf"[*:1]", text)
    return modified_text

def add_isotope(core):
    core_mol=get_mol(core)
    core_mol_atoms=core_mol.GetAtoms()
    matchDummy = core_mol.GetSubstructMatches(Chem.MolFromSmarts('[#0]'))
    for idx,imatch in enumerate(matchDummy):
        iDummy = imatch[0]
        core_mol_atoms[iDummy].SetIsotope(idx)
        # leftDummy=matchDummy[0][0]
        # rightDummy=matchDummy[1][0]
        
        # core_mol_atoms[leftDummy].SetIsotope(0)
        # core_mol_atoms[rightDummy].SetIsotope(1)

    coreSmi = Chem.MolToSmiles(core_mol, isomericSmiles=True)
    # coreSmi = replace_pattern(coreSmi)
    # print(coreSmi)
    return coreSmi

def get_dummy_negb(atom):
    ''' Get the neighbor index of the dummy atom '''
    negb=atom.GetNeighbors()[0]
    return int(negb.GetIdx())

def remove_hydrogens(smi):
    mol = get_mol(smi)
    mol = Chem.RemoveHs(mol)
    return Chem.MolToSmiles(mol, isomericSmiles=True)

def replace_numbered_placeholders(text):
    """
    将文本中所有[数字:*]格式的字符串替换为*
    
    参数:
        text: 包含[1:*]、[2:*]等格式的文本
    
    返回:
        替换后的文本
    """
    # 正则表达式模式：匹配[数字:*]
    # \d+ 匹配一个或多个数字
    # :\* 匹配:*（*需要转义）
    pattern = r'\[\d+\*\]'
    
    # 替换匹配的模式为*
    replaced_text = re.sub(pattern, '*', text)
    return replaced_text

# coreSmi = core
def get_RGroups(smi,coreSmi):
    try:
        core_mol=Chem.MolFromSmarts(coreSmi)
        core_atoms = core_mol.GetAtoms()
        matchDummy = core_mol.GetSubstructMatches(Chem.MolFromSmarts('[#0]'))
        # print('matchDummy= ', matchDummy)

        dummyInfo = []
        for idummy in matchDummy:
            dummyID = idummy[0]
            isotopeDummy = core_mol.GetAtomWithIdx(dummyID).GetIsotope()
            dummyAtom = core_atoms[dummyID]
            dummyNeighbor = get_dummy_negb(dummyAtom)
            dummyInfo.append({'atomId':dummyID, 'isotopeDummy':isotopeDummy, 'dummyNeighbor': dummyNeighbor, 'R':[]})

        # print(dummyInfo)
        mol=Chem.MolFromSmiles(smi)
        mol=Chem.AddHs(mol)
        smart_mol = Chem.MolFromSmarts(replace_pattern(coreSmi))
        matches = mol.GetSubstructMatches(smart_mol, uniquify=False)
        if len(matches) == 0:
            return None
        # Atoms=mol.GetAtoms()
        # print(match)
        RGroupsList = []
        for match in matches:   ### sometimes more than one match
            RGroups = {}  
            bondList = []
            for idx,idummy in enumerate(dummyInfo):
                dummyID  = idummy['atomId']
                isotopeDummy = idummy['isotopeDummy']
                dummyNeighbor =  idummy['dummyNeighbor']
                atomIdx_dummy = match[dummyID]  ## mathched dummy position
                atomIdx_dummyNeighbor = match[dummyNeighbor]  ## mathched dummy neighbour position
                bond = mol.GetBondBetweenAtoms(atomIdx_dummy, atomIdx_dummyNeighbor)
                bondIdx = bond.GetIdx()
                bondList.append(bondIdx)
                dummyInfo[idx]['atomIdx_dummy'] = atomIdx_dummy


            bondList = sorted(bondList)
            # print(bondList)
            nm = Chem.FragmentOnBonds(mol,bondList)
            Mols=Chem.GetMolFrags(nm,asMols=False)
            # print(Mols)
            # png,svg = show_mols([coreSmi])
            # # display(png)
            
            if len(Mols) == len(dummyInfo) + 1:
                for idx,idummy in enumerate(dummyInfo):
                    # print(idx,idummy)
                    atomIdx_dummy  = idummy['atomIdx_dummy']
                    isotopeDummy = idummy['isotopeDummy']
                    # atomIdx_dummyNeighbor =  idummy['atomIdx_dummyNeighbor']
                    for iMol in Mols:
                        # print(iMol)
                        if atomIdx_dummy in iMol:
                    
                            # 生成碎片的SMILES（保留原子间的键）
                            fragment_smiles = Chem.MolFragmentToSmiles(
                                nm, 
                                atomsToUse=iMol, 
                                isomericSmiles=True  # 保留异构信息
                            )
                            # print('fragment_smiles=  ', fragment_smiles)
                            fragment_smiles = remove_hydrogens(fragment_smiles)
                            RGroups[isotopeDummy] = replace_numbered_placeholders(fragment_smiles)
            RGroupsList.append(RGroups)
        return RGroupsList
    except Exception as e:
        print(e)
        return []

def generate_all_excel_columns(max_column=16384):
    """
    生成所有Excel列字母（从A到XFD）
    
    参数:
        max_column: 最大列号，Excel最大支持16384（XFD）
    
    返回:
        包含所有列字母的列表
    """
    if max_column < 1 or max_column > 16384:
        raise ValueError("最大列号必须在1到16384之间")
    
    columns = []
    for num in range(1, max_column + 1):
        columns.append(get_excel_column_letter(num))
    return columns

def get_excel_column_letter(column_number):
    """
    将列号转换为Excel列字母（如1→A, 27→AA）
    
    参数:
        column_number: 列号（整数，从1开始）
    
    返回:
        对应的Excel列字母字符串
    """
    if column_number < 1:
        raise ValueError("列号必须大于等于1")
    
    letters = []
    while column_number > 0:
        # 调整为0-25范围（A-Z对应0-25）
        column_number -= 1
        # 获取当前字符（A的ASCII码是65）
        letters.append(chr(column_number % 26 + 65))
        # 处理下一位
        column_number = column_number // 26
    
    # 反转得到正确顺序
    return ''.join(reversed(letters))

def set_isotope(mol, isotope):
    '''  set isotope for the first matched dummy atom  '''
    mol = copy.deepcopy(mol)
    mol_atoms=mol.GetAtoms()
    matchDummy = mol.GetSubstructMatches(Chem.MolFromSmarts('[#0]'))
    if len(matchDummy) > 0:
        mol_atoms[matchDummy[0][0]].SetIsotope(isotope)
    return mol

def mol2img(mol, imgPath):
    mol = get_mol(mol)
    img = Draw.MolToImage(mol, size=[300, 300])
    img.save(tmpImgPath.joinpath(imgPath))
    img=Image(tmpImgPath.joinpath(imgPath))
    img.width = 100
    img.height = 100
    return img

def connect_core_Rs(iRDicts, core='', return_type='smiles'):
    ''' Connect R groups to the core
    '''
    core_mol=get_mol(core)
    core_mol = copy.deepcopy(core_mol)  ### To protect the inpur molecule object
    # png,svg = show_mols([core_mol])
    # display(svg)

    mols = [core_mol]
    for iR in iRDicts:
        R = iR['R']
        isotope = iR['isotope']
        r_mol=Chem.MolFromSmiles(R)  ### the isotope of dummy atom is zero
        r_mol=set_isotope(r_mol, isotope)
        # png,svg = show_mols([r_mol])
        # display(svg)
        mols.append(r_mol)

    combined_mol = reduce(Chem.CombineMols, mols)
    match = combined_mol.GetSubstructMatches(Chem.MolFromSmarts('[#0]')) ### detect the dummy atoms
    isotopeDict = {}
    dnmmyList = []  ## keep all the dummy atoms
    for imatch in match:
        atm_idx=imatch[0]
        isotope=combined_mol.GetAtomWithIdx(atm_idx).GetIsotope()
        dnmmyList.append(atm_idx)  
        dummyNeighb = get_dummy_negb(combined_mol.GetAtomWithIdx(atm_idx))
        if isotope not in isotopeDict.keys():
            isotopeDict[isotope]=[dummyNeighb]
        else:
            isotopeDict[isotope].append(dummyNeighb)
    # print('isotopeDict= ', isotopeDict)

    edcombo = Chem.EditableMol(combined_mol)
    for isotope, idx_list in isotopeDict.items():
        # print('idx_list= ', idx_list)
        if len(idx_list) == 2:  ## two dummy atoms with same isotope
            edcombo.AddBond(idx_list[0],idx_list[1],order=Chem.rdchem.BondType.SINGLE)  
    
    for idummy in sorted(dnmmyList, reverse=True): # 从高索引到低索引删除原子（避免索引偏移）
        edcombo.RemoveAtom(idummy)
    

    combo = edcombo.GetMol()
    # png,svg = show_mols([combo])
    # display(svg)

    ''' Replace dummy atom with hydrogen '''
    products = Chem.ReplaceSubstructs(combo,Chem.MolFromSmarts('[#0]'),Chem.MolFromSmarts('[#1]'),replaceAll=True)
    combo=products[0]
    products = Chem.ReplaceSubstructs(combo,Chem.MolFromSmarts('[#50]'),Chem.MolFromSmarts('[#0]'),replaceAll=True)
    combo=products[0]
    combo_smi=Chem.MolToSmiles(combo)  ### move the hydrogen
    combo=Chem.MolFromSmiles(combo_smi) 
    # combo=Chem.RemoveHs(combo)
    if return_type=='mol':
        return combo
    if return_type=='smiles':
        combo_smi=Chem.MolToSmiles(combo)
        return combo_smi

def get_activity_info(frag, df_act,  actCols=[], smilesCol='smiles', onlyValue=False):
    '''   Get the activity value of a fragment according molecules   '''
    # try:
    if 1:
        if isinstance(df_act, str):
            df_act = pd.read_csv(df_act)
        else:
            df_act=df_act.copy()
        df_act['similarity']=df_act.apply(lambda x: compute_sim(x[smilesCol], frag), axis=1)
        
        df_match=df_act[df_act['similarity']>0.99]
        if len(df_match) == 0:
            return ''
        
        df_match = df_match[actCols]
        meanStdMedianDict = {'mean':{}, 'std':{}, 'median':{}, 'min':{}, 'max':{}} ##

        for iactCol in actCols:
            dfMatchTmp = df_match.copy()
            
            dfMatchTmp = dfMatchTmp.dropna(subset=[iactCol])
            singleValue = False
            if len(dfMatchTmp)==1:
                singleValue = True
                # meanStdMedianDict['mean'][iactCol]=float(dfMatchTmp[0][iactCol])
                # meanStdMedianDict['std'][iactCol]=float(dfMatchTmp[0][iactCol])
                # meanStdMedianDict['median'][iactCol]=float(dfMatchTmp[0][iactCol])
                # meanStdMedianDict['min'][iactCol]=float(dfMatchTmp[0][iactCol])
                # meanStdMedianDict['max'][iactCol]=float(dfMatchTmp[0][iactCol])

                meanStdMedianDict['mean'][iactCol]=float(dfMatchTmp.iloc[0][iactCol])
                meanStdMedianDict['std'][iactCol]=float(dfMatchTmp.iloc[0][iactCol])
                meanStdMedianDict['median'][iactCol]=float(dfMatchTmp.iloc[0][iactCol])
                meanStdMedianDict['min'][iactCol]=float(dfMatchTmp.iloc[0][iactCol])
                meanStdMedianDict['max'][iactCol]=float(dfMatchTmp.iloc[0][iactCol])

            if len(dfMatchTmp)>1:
                meanStdMedianDict['mean'][iactCol]=float(dfMatchTmp[iactCol].mean())
                meanStdMedianDict['std'][iactCol]=float(dfMatchTmp[iactCol].std())
                meanStdMedianDict['median'][iactCol]=float(dfMatchTmp[iactCol].median())
                meanStdMedianDict['min'][iactCol]=float(dfMatchTmp[iactCol].min())
                meanStdMedianDict['max'][iactCol]=float(dfMatchTmp[iactCol].max())

        actInfo=''
        if onlyValue and len(actCols)==1:
            itarget = actCols[0]
            actInfo =round(float(meanStdMedianDict['mean'][itarget]),1)
            return actInfo
        
        else:
            for itarget in actCols:#["JAK1ToJAK2","JAK1","JAK2"]:
                iInfo = f"{itarget}"
                for imetric in meanStdMedianDict.keys(): #['mean','std','median']:
                    if itarget in meanStdMedianDict[imetric].keys():
                        iInfo += f" {round(float(meanStdMedianDict[imetric][itarget]), 1)}"
                    if singleValue:
                        break  ## only one value is necessay

                if iInfo != f"{itarget}":
                    iInfo+=" | \n"
                    actInfo+=iInfo

            return actInfo
    
    # except Exception as e:
    #     print(e)
    #     return ''
    
def get_activity_info_dict(frag='', df_act='',  onlyValue=0, smilesCol='smiles',  actCols=[]):
    actInfo = get_activity_info(frag, df_act, actCols=actCols, onlyValue=onlyValue, smilesCol=smilesCol)
    print(frag,  actInfo)
    return {frag: actInfo}

def df_valid(df, row_smi='SMILES'):
    valid_idx=[idx for idx, row in df.iterrows() if canonic_smiles(row[row_smi])!=None]
    df_valid=df.loc[valid_idx]
    return df_valid

if __name__ == "__main__":
    import argparse
    class argNameSpace():
        def __init__(self) -> None:
            self.A=""

    def get_parser():
        parser = argparse.ArgumentParser()
        parser.add_argument("--csvFile", help="the name of csv file", required=True, default="Test/input.csv")
        parser.add_argument("--saveFile", help="file to save the results", required=False, default=None)
        parser.add_argument("--smilesCol", help="the column name of SMILES", required=False, default='smiles')
        parser.add_argument("--core",  help="maxmium images", required=False, type=str, default="[*]C(c1c(OC)cc(C)c(Sc2sc(NC([*])=O)nc2)c1)=O")
        parser.add_argument("--actCols", nargs='+', help="the column name of activity", required=False, default=['ITK','JAK3','selectivity'])
        args = parser.parse_args()
        return args

    test=0
    if test>0:
        args=argNameSpace
        args.csvFile = "Test/input.csv"
        args.smilesCol = "smiles"
        # args.core = "[*]C(c1c(OC)cc(C)c(Sc2sc(NC([*])=O)nc2)c1)=O"
        args.core = "[*]C(c1c(OC)cc(C)c(Sc2sc(NC(c3csnc3)=O)nc2)c1)=O"
        args.actCols = ['ITK','JAK3','selectivity']

    else:
        args = get_parser()
        print(args)

    csvFile = args.csvFile
    # smi = 'C=CC(N1CCCN(C[C@H]1C)C(c2c(OC)cc(C)c(Sc3sc(NC=O)nc3)c2)=O)=O'
    core = args.core
    smiCol = args.smilesCol
    actCols = args.actCols

    coreSmi = add_isotope(core)   ## add isotope for differeciate the dummy atom
    print('coreSmi=  ', coreSmi)
    get_RGroups_P = partial(get_RGroups, coreSmi=coreSmi)

    dfInput = pd.read_csv(csvFile)
    
    # RDict = mapper(50)(get_RGroups_P, dfInput[smiCol].tolist())
    RDict = []
    for ismi in dfInput[smiCol]:
        RDict.extend(get_RGroups_P(ismi))
    RDicts = [i for i in RDict if i is not None]
    dfR = pd.DataFrame.from_dict(RDicts)

    wb = openpyxl.Workbook()
    # ws = wb.create_sheet('Sheet1')
    ws = wb.active
    tmpImgPath=Path('./images_png')
    if tmpImgPath.exists():
        shutil.rmtree(tmpImgPath)
    tmpImgPath.mkdir(exist_ok=True, parents=True)

    ws.row_dimensions[1].height = 90
    ws.row_dimensions[2].height = 90
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25

    '''  Add core info to the first line   '''
    ws.cell(1,1).value = 'Core: '
    ws.cell(1,2).value = coreSmi
    ws.cell(1,3).value = 'Activity: '
    ws.cell(1,4).value = str(actCols)
    coreImg = mol2img(coreSmi, 'core.png')
    # coreImg.anchor = ws.cell(row=1, column=2).coordinate # col and row are zero-based
    ws.add_image(coreImg, f"{get_excel_column_letter(2)}{1}")

    if len(RDicts[0]) == 1:  ##  1D SAR table
        molLocation = {} 

        R0Counter = Counter(dfR[0])
        R0Counter = sorted(R0Counter.items(), key=lambda x: x[1], reverse=True)
        for idx0, iR0 in enumerate(R0Counter):
            iR0 = iR0[0]
            
            imgRow = idx0 + 3   ## begin from +2, the first column is for core the second column is for R1
            ws.cell(imgRow, 1).value = iR0   ### begin from +2, the first column is for core
            R0Img = mol2img(iR0, f'R0_{idx0}.png')
            ws.add_image(R0Img, f"{get_excel_column_letter(1)}{imgRow}")
            ws.row_dimensions[imgRow].height = 90

            iRDictList = [{'R':iR0, 'isotope':0}]
            smiComplete = connect_core_Rs(iRDictList, coreSmi, return_type='smiles')  ## connect core and R groups

            molLocation[smiComplete] = [imgRow, 1]

        print(molLocation)
        for idx, icol in enumerate(actCols):
            ws.column_dimensions[get_excel_column_letter(1+1+idx)].width = 25
            ws.cell(2, 1+1+idx).value = icol   ### begin from +2, the first column is for core
            get_activity_info_dict_p = partial(get_activity_info_dict, df_act=csvFile,  actCols=[icol],  onlyValue = 1, smilesCol='smiles')
            actDictList = mapper(50)(get_activity_info_dict_p, molLocation.keys())
            mergedActInfo = {}

            for d in actDictList:
                # Update merged dict with key-value pairs from current dict
                mergedActInfo.update(d)

            for ikey,ivalue in molLocation.items():
                ws.cell(ivalue[0], ivalue[1]+1+idx).value = mergedActInfo[ikey]
                    

    if len(RDicts[0]) == 2:  ##  2D SAR table
        R0Count = len(set(dfR[0]))
        R1Count = len(set(dfR[1]))

        if R0Count > R1Count:
            R0Counter = Counter(dfR[0])
            R0Counter = sorted(R0Counter.items(), key=lambda x: x[1], reverse=True)

            R1Counter = Counter(dfR[1])
            R1Counter = sorted(R1Counter.items(), key=lambda x: x[1], reverse=True)

        if R0Count < R1Count:   ##  make the longer one as the y aixs
            R0Counter = Counter(dfR[1])
            R0Counter = sorted(R0Counter.items(), key=lambda x: x[1], reverse=True)
            
            R1Counter = Counter(dfR[0])
            R1Counter = sorted(R1Counter.items(), key=lambda x: x[1], reverse=True)
            
        molLocation = {}

        for idx0, iR0 in enumerate(R0Counter):
            iR0 = iR0[0]
            if 1:   ## R0 as Y-axis;  R1 as X-axis， the larger one is Y-axis
                imgRow = idx0+3   ## begin from +2, the first column is for core the second column is for R1
                ws.cell(imgRow, 1).value = iR0   ### begin from +2, the first column is for core
                R0Img = mol2img(iR0, f'R0_{idx0}.png')
                ws.add_image(R0Img, f"{get_excel_column_letter(1)}{imgRow}")
                ws.row_dimensions[imgRow].height = 90
                
            else: ## R1 as Y-axis;  R0 as X-axis
                pass

            for idx1, iR1 in enumerate(R1Counter):
                # print(f"R1: {iR1}")
                iR1 = iR1[0]
                if 1:   ## R0 as Y-axis;  R1 as X-axis， the larger one is Y-axis
                    if idx0 == 0:  ## Only need add R0 once in the first row
                        imgCol = idx1+2  ## +1 for not zero-index; +1 for not R0
                        ws.cell(2, imgCol).value = iR1   ### begin from +1, the first column is for core
                        R1Img = mol2img(iR1, f'R1_{idx1}.png')
                        ws.add_image(R1Img, f"{get_excel_column_letter(imgCol)}{idx0+2}")
                        ws.column_dimensions[get_excel_column_letter(imgCol)].width = 25

                else: ## R1 as Y-axis;  R0 as X-axis
                    pass
                
                if R0Count > R1Count:
                    iRDictList = [{'R':iR0, 'isotope':0}, {'R':iR1, 'isotope':1}]
                if R0Count < R1Count:
                    iRDictList = [{'R':iR0, 'isotope':1}, {'R':iR1, 'isotope':0}]

                smiComplete = connect_core_Rs(iRDictList, coreSmi, return_type='smiles')  ## connect core and R groups
                
                # actInfo = get_activity_info(smiComplete, dfInput,  actCols=['ITK','JAK3','selectivity'], smilesCol='smiles')
                if 1: 
                    molLocation[smiComplete] = [idx0+3 , idx1+2]



        print(molLocation)
        get_activity_info_dict_p = partial(get_activity_info_dict, df_act=csvFile, onlyValue = 1,  actCols=actCols,  smilesCol='smiles')
        actDictList = mapper(50)(get_activity_info_dict_p, molLocation.keys())
        mergedActInfo = {}

        for d in actDictList:
            # Update merged dict with key-value pairs from current dict
            mergedActInfo.update(d)

        for ikey, ivalue in molLocation.items():
            ws.cell(ivalue[0], ivalue[1]).value = mergedActInfo[ikey]

    if not args.saveFile:            
        wb.save(csvFile.replace('.csv', '.xlsx'))
    else:
        wb.save(args.saveFile)
    wb.close()
